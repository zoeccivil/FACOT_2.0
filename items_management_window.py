from __future__ import annotations

import re
import sqlite3
from typing import Optional, Tuple, List, Dict

from PyQt6.QtCore import Qt
from PyQt6.QtGui import QDoubleValidator, QAction, QFontMetrics
from PyQt6.QtWidgets import (
    QDialog, QVBoxLayout, QHBoxLayout, QLabel, QLineEdit, QPushButton, QComboBox,
    QTableWidget, QTableWidgetItem, QMessageBox, QHeaderView, QWidget, QFormLayout,
    QTextEdit, QSpinBox, QFileDialog, QSizePolicy, QProgressDialog, QMenu
)
from PyQt6.QtWidgets import QWidget
# Excel
try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.worksheet.datavalidation import DataValidation
except Exception:
    Workbook = None
    load_workbook = None
    DataValidation = None

import facot_config  # Ruta de BD desde la app

CODE_PAD = 4  # ABC0001

# -------------------- Utilidades de BD y migraciones -------------------- #
def get_db_path() -> str:
    return facot_config.get_db_path() or ""

def _column_exists(conn: sqlite3.Connection, table: str, column: str) -> bool:
    cur = conn.execute(f"PRAGMA table_info({table})")
    return any(row[1] == column for row in cur.fetchall())

def ensure_items_schema(db_path: str):
    if not db_path:
        return
    with sqlite3.connect(db_path) as conn:
        conn.execute("PRAGMA foreign_keys = ON;")
        # Tablas base
        conn.executescript("""
            CREATE TABLE IF NOT EXISTS categories (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT NOT NULL UNIQUE
            );
            CREATE TABLE IF NOT EXISTS items (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                code TEXT NOT NULL UNIQUE,
                name TEXT NOT NULL,
                unit TEXT NOT NULL,
                cost REAL NOT NULL DEFAULT 0,
                price REAL NOT NULL DEFAULT 0,
                category_id INTEGER,
                description TEXT,
                FOREIGN KEY (category_id) REFERENCES categories(id)
            );
            CREATE INDEX IF NOT EXISTS idx_items_code ON items(code);
            CREATE INDEX IF NOT EXISTS idx_items_category ON items(category_id);
        """)
        # Migraciones: agregar code_prefix, next_seq, description a categories
        if not _column_exists(conn, "categories", "code_prefix"):
            conn.execute("ALTER TABLE categories ADD COLUMN code_prefix TEXT;")
        if not _column_exists(conn, "categories", "next_seq"):
            conn.execute("ALTER TABLE categories ADD COLUMN next_seq INTEGER NOT NULL DEFAULT 1;")
        if not _column_exists(conn, "categories", "description"):
            conn.execute("ALTER TABLE categories ADD COLUMN description TEXT;")
        try:
            conn.execute("CREATE UNIQUE INDEX IF NOT EXISTS idx_categories_code_prefix ON categories(code_prefix);")
        except Exception:
            pass
        _backfill_category_meta(conn)

def _slug_letters(text: str, n: int = 3) -> str:
    letters = re.findall(r"[A-Za-z]", text)
    if not letters:
        return "CAT"
    return "".join(letters[:n]).upper().ljust(n, "X")

def _max_seq_for_prefix(conn: sqlite3.Connection, prefix: str) -> int:
    if not prefix:
        return 0
    cur = conn.execute(
        "SELECT MAX(CAST(SUBSTR(code, ?) AS INTEGER)) FROM items WHERE code LIKE ?",
        (len(prefix) + 1, f"{prefix}%"),
    )
    row = cur.fetchone()
    try:
        return int(row[0]) if row and row[0] is not None else 0
    except Exception:
        return 0

def _backfill_category_meta(conn: sqlite3.Connection):
    cur = conn.execute("SELECT id, name, code_prefix, next_seq FROM categories")
    for cid, name, prefix, next_seq in cur.fetchall():
        updated = False
        if not prefix:
            base = _slug_letters(name)
            candidate = base
            i = 1
            while True:
                row = conn.execute("SELECT id FROM categories WHERE code_prefix = ? AND id != ?", (candidate, cid)).fetchone()
                if not row:
                    break
                i += 1
                candidate = f"{base[:-1]}{i%10}"
            prefix = candidate
            conn.execute("UPDATE categories SET code_prefix=? WHERE id=?", (prefix, cid))
            updated = True
        if not next_seq or next_seq < 1:
            next_seq = 1
        max_seq = _max_seq_for_prefix(conn, prefix)
        if max_seq >= next_seq:
            conn.execute("UPDATE categories SET next_seq=? WHERE id=?", (max_seq + 1, cid))
            updated = True
        if updated:
            conn.commit()

def get_next_code(db_path: str, category_id: int) -> Optional[str]:
    if not db_path or not category_id:
        return None
    with sqlite3.connect(db_path) as conn:
        row = conn.execute("SELECT code_prefix, next_seq FROM categories WHERE id=?", (category_id,)).fetchone()
        if not row:
            return None
        prefix, seq = row
        if not prefix:
            prefix = "CAT"
        seq = int(seq or 1)
        return f"{prefix}{seq:0{CODE_PAD}d}"

def bump_category_seq(db_path: str, category_id: int):
    if not db_path or not category_id:
        return
    with sqlite3.connect(db_path) as conn:
        conn.execute("UPDATE categories SET next_seq = IFNULL(next_seq,1) + 1 WHERE id=?", (category_id,))
        conn.commit()

# -------------------- Diálogo de Categoría -------------------- #
class CategoryDialog(QDialog):
    def __init__(self, db_path: str, category: Optional[Tuple]=None, parent: Optional[QWidget]=None):
        super().__init__(parent)
        self.setWindowTitle("Categoría")
        self.db_path = db_path
        self.category = category  # (id, name, code_prefix, next_seq, description)
        self._build_ui()
        if category:
            self._load(category)

    def _build_ui(self):
        self.setMinimumWidth(420)
        lay = QVBoxLayout(self)
        form = QFormLayout()
        self.name_edit = QLineEdit()
        self.prefix_edit = QLineEdit()
        self.seq_spin = QSpinBox(); self.seq_spin.setMinimum(1); self.seq_spin.setMaximum(999999)
        self.desc_edit = QTextEdit(); self.desc_edit.setPlaceholderText("Descripción (opcional)")
        form.addRow("Nombre:", self.name_edit)
        form.addRow("Prefijo de código:", self.prefix_edit)
        form.addRow("Siguiente secuencia:", self.seq_spin)
        form.addRow("Descripción:", self.desc_edit)
        lay.addLayout(form)

        self.name_edit.textChanged.connect(self._on_name_change)
        self.prefix_edit.textEdited.connect(lambda _: None)

        btns = QHBoxLayout()
        ok = QPushButton("Guardar"); ok.clicked.connect(self.accept)
        cancel = QPushButton("Cancelar"); cancel.clicked.connect(self.reject)
        btns.addWidget(ok); btns.addWidget(cancel)
        lay.addLayout(btns)

    def _on_name_change(self, text: str):
        if not self.category or not self.prefix_edit.text():
            self.prefix_edit.setText(_slug_letters(text, 3))

    def _load(self, cat):
        _, name, prefix, next_seq, desc = cat
        self.name_edit.setText(name or "")
        self.prefix_edit.setText(prefix or _slug_letters(name or ""))
        self.seq_spin.setValue(int(next_seq or 1))
        self.desc_edit.setPlainText(desc or "")

    def get_data(self) -> Optional[Tuple[str, str, int, str]]:
        name = self.name_edit.text().strip()
        prefix = self.prefix_edit.text().strip().upper()
        seq = int(self.seq_spin.value())
        desc = self.desc_edit.toPlainText().strip()
        if not name:
            QMessageBox.warning(self, "Validación", "El nombre es obligatorio.")
            return None
        if not re.fullmatch(r"[A-Z0-9]{2,6}", prefix or ""):
            QMessageBox.warning(self, "Validación", "El prefijo debe tener entre 2 y 6 caracteres alfanuméricos en mayúscula.")
            return None
        with sqlite3.connect(self.db_path) as conn:
            if self.category:
                cid = self.category[0]
                row = conn.execute("SELECT id FROM categories WHERE code_prefix = ? AND id != ?", (prefix, cid)).fetchone()
            else:
                row = conn.execute("SELECT id FROM categories WHERE code_prefix = ?", (prefix,)).fetchone()
            if row:
                QMessageBox.warning(self, "Prefijo duplicado", "El prefijo ya existe en otra categoría.")
                return None
        return name, prefix, seq, desc

# -------------------- Diálogo de Ítem -------------------- #
class ItemDialog(QDialog):
    def __init__(self, db_path: str, categories: list[Tuple[int, str, str, int]], item: Optional[dict]=None, parent: Optional[QWidget]=None):
        super().__init__(parent)
        self.setWindowTitle("Ítem")
        self.db_path = db_path
        self.categories = categories  # [(id, name, prefix, next_seq)]
        self.item = item
        self._build_ui()
        if item:
            self._load(item)
        else:
            self._update_code_preview()

    def _build_ui(self):
        self.setMinimumWidth(520)
        lay = QVBoxLayout(self)
        form = QFormLayout()

        self.cat_combo = QComboBox()
        self.cat_combo.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Fixed)
        for cid, name, prefix, next_seq in self.categories:
            self.cat_combo.addItem(f"{name} ({prefix})", cid)
        self.cat_combo.currentIndexChanged.connect(self._update_code_preview)

        self.auto_code = QComboBox()
        self.auto_code.addItems(["Automático", "Manual"])
        self.auto_code.currentIndexChanged.connect(self._on_code_mode_change)

        self.code_edit = QLineEdit()
        self.code_edit.setPlaceholderText("Código (si Manual)")
        self.code_edit.setEnabled(False)

        self.name_edit = QLineEdit()
        self.unit_combo = QComboBox()
        self.unit_combo.setEditable(True)
        self.unit_combo.addItems(["UND", "SERV", "HR", "M", "M2", "M3", "KG", "LT", "PAQ"])

        self.cost_edit = QLineEdit(); self.cost_edit.setValidator(QDoubleValidator(0.0, 1e12, 4))
        self.price_edit = QLineEdit(); self.price_edit.setValidator(QDoubleValidator(0.0, 1e12, 4))
        self.desc_edit = QTextEdit()

        self.next_code_label = QLabel("Siguiente código: -")

        code_row = QHBoxLayout()
        code_row.addWidget(self.auto_code)
        code_row.addWidget(self.code_edit)
        code_cont = QWidget(); code_cont.setLayout(code_row)

        form.addRow("Categoría:", self.cat_combo)
        form.addRow("Código:", code_cont)
        form.addRow("Sugerencia:", self.next_code_label)
        form.addRow("Nombre:", self.name_edit)
        form.addRow("Unidad:", self.unit_combo)
        form.addRow("Costo:", self.cost_edit)
        form.addRow("Precio:", self.price_edit)
        form.addRow("Descripción:", self.desc_edit)
        lay.addLayout(form)

        btns = QHBoxLayout()
        ok = QPushButton("Guardar"); ok.clicked.connect(self.accept)
        cancel = QPushButton("Cancelar"); cancel.clicked.connect(self.reject)
        btns.addWidget(ok); btns.addWidget(cancel)
        lay.addLayout(btns)

    def _on_code_mode_change(self, idx: int):
        manual = (self.auto_code.currentText() == "Manual")
        self.code_edit.setEnabled(manual)
        self._update_code_preview()

    def _update_code_preview(self):
        cid = self.cat_combo.currentData()
        suggestion = get_next_code(self.db_path, cid) or "-"
        self.next_code_label.setText(f"Siguiente código: {suggestion}")
        if self.auto_code.currentText() == "Automático":
            self.code_edit.setText("")

    def _load(self, item: dict):
        idx = self.cat_combo.findData(item["category_id"])
        if idx >= 0:
            self.cat_combo.setCurrentIndex(idx)
        self.auto_code.setCurrentText("Manual")
        self.code_edit.setEnabled(True)
        self.code_edit.setText(item["code"])
        self.name_edit.setText(item["name"])
        uidx = self.unit_combo.findText(item["unit"])
        if uidx >= 0:
            self.unit_combo.setCurrentIndex(uidx)
        else:
            self.unit_combo.setEditText(item["unit"])
        self.cost_edit.setText(str(item["cost"]))
        self.price_edit.setText(str(item["price"]))
        self.desc_edit.setPlainText(item.get("description", ""))

    def get_data(self) -> Optional[dict]:
        cid = int(self.cat_combo.currentData() or 0)
        if not cid:
            QMessageBox.warning(self, "Validación", "Seleccione una categoría.")
            return None
        name = self.name_edit.text().strip()
        unit = self.unit_combo.currentText().strip().upper()
        try:
            cost = float(self.cost_edit.text().replace(",", ".") or 0)
            price = float(self.price_edit.text().replace(",", ".") or 0)
        except ValueError:
            QMessageBox.warning(self, "Validación", "Costo y Precio deben ser numéricos.")
            return None
        manual = (self.auto_code.currentText() == "Manual")
        code = self.code_edit.text().strip().upper()
        if manual and not code:
            QMessageBox.warning(self, "Validación", "Ingrese un código o use Automático.")
            return None
        return {
            "category_id": cid,
            "code": code if manual else None,
            "name": name,
            "unit": unit,
            "cost": cost,
            "price": price,
            "description": self.desc_edit.toPlainText().strip(),
            "manual_code": manual
        }

# -------------------- Importador Excel (con vista previa y progreso) -------------------- #
class ExcelImportDialog(QDialog):
    """
    Importa Categorías e Ítems desde una plantilla Excel.
    Hojas esperadas:
    - Categorias: name, code_prefix, next_seq, description
    - Items: code(optional), name, unit, cost, price, category_prefix, description
    """
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Importar desde Excel")
        self.setMinimumWidth(760)
        self._build_ui()

    def _build_ui(self):
        lay = QVBoxLayout(self)
        row = QHBoxLayout()
        row.addWidget(QLabel("Archivo Excel:"))
        self.path_edit = QLineEdit()
        self.path_edit.setPlaceholderText("Selecciona el archivo .xlsx a importar")
        btn_browse = QPushButton("Examinar…")
        btn_browse.clicked.connect(self._browse)
        row.addWidget(self.path_edit, stretch=1)
        row.addWidget(btn_browse)
        lay.addLayout(row)

        # Vista previa de validación
        self.preview = QTextEdit()
        self.preview.setReadOnly(True)
        self.preview.setPlaceholderText("Aquí verás un resumen de lo que se creará/actualizará antes de importar…")
        lay.addWidget(self.preview, stretch=1)

        # Botonera
        bot = QHBoxLayout()
        self.btn_preview = QPushButton("Previsualizar")
        self.btn_preview.clicked.connect(self._preview)
        self.btn_import = QPushButton("Importar")
        self.btn_import.clicked.connect(self._do_import)
        self.btn_cancel = QPushButton("Cancelar")
        self.btn_cancel.clicked.connect(self.reject)
        bot.addStretch(1)
        bot.addWidget(self.btn_preview)
        bot.addWidget(self.btn_import)
        bot.addWidget(self.btn_cancel)
        lay.addLayout(bot)

    def _browse(self):
        fn, _ = QFileDialog.getOpenFileName(self, "Selecciona Excel", "", "Excel (*.xlsx)")
        if fn:
            self.path_edit.setText(fn)

    # --- Previsualización (no escribe en la BD) ---
    def _preview(self):
        if load_workbook is None:
            QMessageBox.critical(self, "Dependencia faltante", "Instala openpyxl: pip install openpyxl")
            return
        path = self.path_edit.text().strip()
        if not path:
            QMessageBox.warning(self, "Archivo", "Selecciona un archivo .xlsx")
            return

        db = get_db_path()
        ensure_items_schema(db)

        try:
            wb = load_workbook(path, data_only=True)
        except Exception as e:
            QMessageBox.critical(self, "Excel", f"No se pudo abrir el archivo:\n{e}")
            return

        # Cargar estado actual
        with sqlite3.connect(db) as conn:
            cur = conn.cursor()
            cats_db: Dict[str, Tuple[int, str, int, str]] = {}  # prefix -> (id,name,next_seq,desc)
            for cid, name, prefix, next_seq, desc in cur.execute("SELECT id, name, IFNULL(code_prefix,''), IFNULL(next_seq,1), IFNULL(description,'') FROM categories"):
                cats_db[prefix.upper()] = (cid, name, next_seq, desc)
            items_db_codes = {row[0] for row in cur.execute("SELECT code FROM items")}

        created_cats = updated_cats = 0
        created_items = updated_items = auto_coded = 0
        errors: List[str] = []

        # Categorías
        if "Categorias" in wb.sheetnames:
            sh = wb["Categorias"]
            headers = [str(sh.cell(row=1, column=c).value or "").strip().lower() for c in range(1, sh.max_column+1)]
            def idx(h): 
                return (headers.index(h) + 1) if h in headers else -1
            req = ["name", "code_prefix", "next_seq", "description"]
            miss = [r for r in req if r not in headers]
            if miss:
                errors.append(f"Hoja Categorias: faltan columnas {', '.join(miss)}")
            else:
                for r in range(2, sh.max_row+1):
                    name = (sh.cell(r, idx("name")).value or "").strip()
                    prefix = (sh.cell(r, idx("code_prefix")).value or "").strip().upper()
                    next_seq = sh.cell(r, idx("next_seq")).value
                    if not name and not prefix:
                        continue
                    if not name or not prefix:
                        errors.append(f"Categorías fila {r}: nombre y code_prefix son requeridos"); continue
                    try:
                        next_seq = int(next_seq or 1)
                    except Exception:
                        next_seq = 1
                    if prefix in cats_db:
                        updated_cats += 1
                    else:
                        created_cats += 1
        else:
            errors.append("No se encontró la hoja 'Categorias' (se puede importar ítems si ya existen categorías).")

        # Items
        if "Items" not in wb.sheetnames:
            errors.append("No se encontró la hoja 'Items'.")
        else:
            sh = wb["Items"]
            headers = [str(sh.cell(row=1, column=c).value or "").strip().lower() for c in range(1, sh.max_column+1)]
            def idx(h): 
                return (headers.index(h) + 1) if h in headers else -1
            req = ["name", "unit", "cost", "price", "category_prefix", "description"]
            miss = [r for r in req if r not in headers]
            if miss:
                errors.append(f"Hoja Items: faltan columnas {', '.join(miss)}")
            else:
                for r in range(2, sh.max_row+1):
                    code = (sh.cell(r, idx("code(optional)")).value or "").strip().upper() if "code(optional)" in headers else ""
                    name = (sh.cell(r, idx("name")).value or "").strip()
                    unit = (sh.cell(r, idx("unit")).value or "").strip().upper()
                    cat_prefix = (sh.cell(r, idx("category_prefix")).value or "").strip().upper()
                    if not name and not unit and not cat_prefix and code == "":
                        continue
                    if not cat_prefix:
                        errors.append(f"Ítems fila {r}: category_prefix requerido"); continue
                    if code:
                        if code in items_db_codes:
                            updated_items += 1
                        else:
                            created_items += 1
                    else:
                        # sin código -> autogenerado
                        auto_coded += 1
                        created_items += 1

        text = []
        text.append("Previsualización de importación")
        text.append(f"- Categorías a crear: {created_cats}")
        text.append(f"- Categorías a actualizar: {updated_cats}")
        text.append(f"- Ítems a crear: {created_items} (auto-códigos: {auto_coded})")
        text.append(f"- Ítems a actualizar: {updated_items}")
        if errors:
            text.append("\nAdvertencias/Errores:")
            text.extend([f"• {e}" for e in errors[:80]])
            if len(errors) > 80:
                text.append(f"… y {len(errors)-80} más.")
        self.preview.setPlainText("\n".join(text))

    # --- Importación con progreso ---
    def _do_import(self):
        if load_workbook is None:
            QMessageBox.critical(self, "Dependencia faltante", "Instala openpyxl: pip install openpyxl")
            return
        path = self.path_edit.text().strip()
        if not path:
            QMessageBox.warning(self, "Archivo", "Selecciona un archivo .xlsx")
            return

        db = get_db_path()
        ensure_items_schema(db)

        try:
            wb = load_workbook(path, data_only=True)
        except Exception as e:
            QMessageBox.critical(self, "Excel", f"No se pudo abrir el archivo:\n{e}")
            return

        created_cats = updated_cats = 0
        created_items = updated_items = auto_coded = 0
        errors: List[str] = []

        # Contar filas para el progreso
        total_rows = 0
        if "Categorias" in wb.sheetnames:
            total_rows += max(0, wb["Categorias"].max_row - 1)
        if "Items" in wb.sheetnames:
            total_rows += max(0, wb["Items"].max_row - 1)
        progress = QProgressDialog("Importando datos…", "Cancelar", 0, max(1, total_rows), self)
        progress.setWindowModality(Qt.WindowModality.ApplicationModal)
        progress.setMinimumDuration(200)
        step = 0

        # ----- Categorías -----
        if "Categorias" in wb.sheetnames:
            sh = wb["Categorias"]
            headers = [str(sh.cell(row=1, column=c).value or "").strip().lower() for c in range(1, sh.max_column+1)]
            def idx(h): 
                return (headers.index(h) + 1) if h in headers else -1
            req = ["name", "code_prefix", "next_seq", "description"]
            for r in range(2, sh.max_row+1):
                if progress.wasCanceled():
                    break
                name = (sh.cell(r, idx("name")).value or "").strip()
                prefix = (sh.cell(r, idx("code_prefix")).value or "").strip().upper()
                next_seq = sh.cell(r, idx("next_seq")).value
                desc = (sh.cell(r, idx("description")).value or "").strip()
                if not name and not prefix:
                    step += 1; progress.setValue(step); continue
                if not name or not prefix:
                    errors.append(f"Categorías fila {r}: nombre y code_prefix son requeridos"); step += 1; progress.setValue(step); continue
                if not isinstance(next_seq, (int, float)) or int(next_seq) < 1:
                    next_seq = 1
                next_seq = int(next_seq)
                with sqlite3.connect(db) as conn:
                    row = conn.execute("SELECT id FROM categories WHERE code_prefix=?", (prefix,)).fetchone()
                    if row:
                        conn.execute("UPDATE categories SET name=?, next_seq=?, description=? WHERE id=?", (name, next_seq, desc, row[0]))
                        updated_cats += 1
                    else:
                        try:
                            conn.execute("INSERT INTO categories (name, code_prefix, next_seq, description) VALUES (?,?,?,?)",
                                         (name, prefix, next_seq, desc))
                            created_cats += 1
                        except sqlite3.IntegrityError as e:
                            errors.append(f"Categoría prefijo {prefix}: {e}")
                    conn.commit()
                step += 1
                progress.setValue(step)

        # Mapa prefijo->id
        cat_map = {}
        with sqlite3.connect(db) as conn:
            for cid, prefix in conn.execute("SELECT id, code_prefix FROM categories"):
                cat_map[prefix.upper()] = cid

        # ----- Ítems -----
        if "Items" in wb.sheetnames and not progress.wasCanceled():
            sh = wb["Items"]
            headers = [str(sh.cell(row=1, column=c).value or "").strip().lower() for c in range(1, sh.max_column+1)]
            def idx(h): 
                return (headers.index(h) + 1) if h in headers else -1
            with sqlite3.connect(db) as conn:
                conn.execute("PRAGMA foreign_keys = ON;")
                for r in range(2, sh.max_row+1):
                    if progress.wasCanceled():
                        break
                    code = (sh.cell(r, idx("code(optional)")).value or "").strip().upper() if "code(optional)" in headers else ""
                    name = (sh.cell(r, idx("name")).value or "").strip()
                    unit = (sh.cell(r, idx("unit")).value or "").strip().upper()
                    cost = sh.cell(r, idx("cost")).value
                    price = sh.cell(r, idx("price")).value
                    cat_prefix = (sh.cell(r, idx("category_prefix")).value or "").strip().upper()
                    desc = (sh.cell(r, idx("description")).value or "").strip()

                    if not name and not unit and not cat_prefix and code == "":
                        step += 1; progress.setValue(step); continue
                    if not cat_prefix:
                        errors.append(f"Ítems fila {r}: category_prefix requerido"); step += 1; progress.setValue(step); continue
                    cid = cat_map.get(cat_prefix)
                    if not cid:
                        errors.append(f"Ítems fila {r}: categoría con prefijo '{cat_prefix}' no existe"); step += 1; progress.setValue(step); continue
                    try:
                        cost = float(cost or 0)
                        price = float(price or 0)
                    except ValueError:
                        errors.append(f"Ítems fila {r}: cost/price inválidos"); step += 1; progress.setValue(step); continue

                    if code:
                        row = conn.execute("SELECT id FROM items WHERE code=?", (code,)).fetchone()
                        if row:
                            conn.execute("""
                                UPDATE items SET name=?, unit=?, cost=?, price=?, category_id=?, description=? WHERE code=?
                            """, (name, unit, cost, price, cid, desc, code))
                            updated_items += 1
                        else:
                            try:
                                conn.execute("""
                                    INSERT INTO items (code, name, unit, cost, price, category_id, description)
                                    VALUES (?,?,?,?,?,?,?)
                                """, (code, name, unit, cost, price, cid, desc))
                                created_items += 1
                            except sqlite3.IntegrityError as e:
                                errors.append(f"Ítems fila {r}: {e}")
                    else:
                        next_code = get_next_code(get_db_path(), cid)
                        if not next_code:
                            errors.append(f"Ítems fila {r}: no se pudo generar código para prefijo {cat_prefix}")
                            step += 1; progress.setValue(step); continue
                        try:
                            conn.execute("""
                                INSERT INTO items (code, name, unit, cost, price, category_id, description)
                                VALUES (?,?,?,?,?,?,?)
                            """, (next_code, name, unit, cost, price, cid, desc))
                            created_items += 1
                            auto_coded += 1
                            bump_category_seq(get_db_path(), cid)
                        except sqlite3.IntegrityError as e:
                            errors.append(f"Ítems fila {r}: {e}")
                    conn.commit()
                    step += 1
                    progress.setValue(step)

        progress.setValue(total_rows)

        summary = (
            f"Categorías: +{created_cats} creadas, {updated_cats} actualizadas\n"
            f"Ítems: +{created_items} creados, {updated_items} actualizados (auto-códigos: {auto_coded})\n"
        )
        if errors:
            summary += "\nAdvertencias/Errores:\n- " + "\n- ".join(errors[:80])
            if len(errors) > 80:
                summary += f"\n… y {len(errors)-80} más."
        QMessageBox.information(self, "Importación finalizada", summary)
        self.accept()

# -------------------- Generador/Exportador de Excel -------------------- #
def generate_excel_template(save_path: str):
    if Workbook is None or DataValidation is None:
        raise RuntimeError("openpyxl no está instalado. Instala con: pip install openpyxl")
    wb = Workbook()

    # Hoja Categorias
    shc = wb.active
    shc.title = "Categorias"
    shc.append(["name", "code_prefix", "next_seq", "description"])
    shc.append(["Servicios", "SERV", 1, "Servicios generales"])
    shc.append(["Insumos Eléctricos", "ELE", 1, "Material eléctrico"])
    shc.append(["Construcción", "CONST", 1, "Materiales de construcción"])

    # Hoja Items
    shi = wb.create_sheet("Items")
    shi.append(["code(optional)", "name", "unit", "cost", "price", "category_prefix", "description"])
    shi.append(["", "Mano de obra técnica", "SERV", 0, 1500, "SERV", "Servicio por hora"])
    shi.append(["ELE0001", "Cable THHN 12AWG", "M", 25, 45, "ELE", "Cobre"])
    shi.append(["", "Cemento gris", "UND", 300, 450, "CONST", "Saco 42.5kg"])

    # Lista de Unidades
    shu = wb.create_sheet("Unidades")
    units = ["UND", "SERV", "HR", "M", "M2", "M3", "KG", "LT", "PAQ"]
    for u in units:
        shu.append([u])

    # Validaciones de datos (usar rangos como string)
    dv_units = DataValidation(type="list", formula1="=Unidades!$A:$A", allow_blank=True)
    shi.add_data_validation(dv_units)
    dv_units.add("C2:C1048576")  # antes: dv_units.add(shi["C2:C1048576"])

    dv_cat = DataValidation(type="list", formula1="=Categorias!$B:$B", allow_blank=False)
    shi.add_data_validation(dv_cat)
    dv_cat.add("F2:F1048576")    # antes: dv_cat.add(shi["F2:F1048576"])

    # Notas
    notes = wb.create_sheet("Notas")
    notes["A1"] = "Instrucciones"
    notes["A2"] = "- En Categorias: code_prefix en MAYÚSCULAS (2–6 caracteres), next_seq = siguiente correlativo."
    notes["A3"] = "- En Items: si 'code(optional)' se deja vacío, el sistema generará el código usando 'category_prefix' y next_seq."
    notes["A4"] = "- 'unit' tiene validación de lista desde la hoja Unidades. 'category_prefix' tiene validación desde Categorias."
    notes["A5"] = "- Puedes agregar más filas debajo de los ejemplos."

    wb.save(save_path)

def export_current_to_excel(save_path: str):
    if Workbook is None or DataValidation is None:
        raise RuntimeError("openpyxl no está instalado. Instala con: pip install openpyxl")
    wb = Workbook()

    # Categorias
    shc = wb.active
    shc.title = "Categorias"
    shc.append(["name", "code_prefix", "next_seq", "description"])

    db = get_db_path()
    ensure_items_schema(db)
    with sqlite3.connect(db) as conn:
        cur = conn.cursor()
        cats = cur.execute("SELECT name, IFNULL(code_prefix,''), IFNULL(next_seq,1), IFNULL(description,'') FROM categories ORDER BY name").fetchall()
        for row in cats:
            shc.append(list(row))

    # Items
    shi = wb.create_sheet("Items")
    shi.append(["code(optional)", "name", "unit", "cost", "price", "category_prefix", "description"])
    with sqlite3.connect(db) as conn:
        rows = conn.execute("""
            SELECT i.code, i.name, i.unit, i.cost, i.price, IFNULL(c.code_prefix,''), IFNULL(i.description,'')
            FROM items i
            LEFT JOIN categories c ON c.id = i.category_id
            ORDER BY i.id
        """).fetchall()
        for row in rows:
            shi.append(list(row))

    # Unidades
    shu = wb.create_sheet("Unidades")
    units = ["UND", "SERV", "HR", "M", "M2", "M3", "KG", "LT", "PAQ"]
    for u in units:
        shu.append([u])

    # Validaciones (usar rangos como string)
    dv_units = DataValidation(type="list", formula1="=Unidades!$A:$A", allow_blank=True)
    shi.add_data_validation(dv_units)
    dv_units.add("C2:C1048576")  # antes: dv_units.add(shi["C2:C1048576"])

    dv_cat = DataValidation(type="list", formula1="=Categorias!$B:$B", allow_blank=False)
    shi.add_data_validation(dv_cat)
    dv_cat.add("F2:F1048576")    # antes: dv_cat.add(shi["F2:F1048576"])

    notes = wb.create_sheet("Notas")
    notes["A1"] = "Exportado desde la aplicación"
    notes["A2"] = "Puedes editar y reimportar este archivo con el diálogo de Importar Excel."
    notes["A3"] = "Si dejas 'code(optional)' vacío, se autogenerará al importar."

    wb.save(save_path)
# -------------------- Ventana principal -------------------- #
class ItemsManagementWindow(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Gestión de Ítems y Categorías")
        # Permitir maximizar y expandir
        self.setWindowFlags(self.windowFlags() | Qt.WindowType.WindowMinMaxButtonsHint)
        self.setSizeGripEnabled(True)
        self.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Expanding)

        self.setMinimumSize(1120, 640)
        ensure_items_schema(get_db_path())
        self._build_ui()
        self._load_categories()
        self._load_items()

    # UI
    def _build_ui(self):
        root = QVBoxLayout(self)

        # Barra de categorías
        cat_bar = QHBoxLayout()
        cat_bar.addWidget(QLabel("Categoría:"))

        self.cat_combo = QComboBox()
        self.cat_combo.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Fixed)
        self.cat_combo.currentIndexChanged.connect(self._on_category_changed)
        self.cat_combo.setMinimumContentsLength(24)
        cat_bar.addWidget(self.cat_combo, stretch=6)

        self.next_code_preview = QLabel("Siguiente código: -")
        cat_bar.addWidget(self.next_code_preview, stretch=3)

        self.btn_new_cat = QPushButton("Nueva")
        self.btn_edit_cat = QPushButton("Editar")
        self.btn_del_cat = QPushButton("Eliminar")
        self.btn_new_cat.clicked.connect(self._new_category)
        self.btn_edit_cat.clicked.connect(self._edit_category)
        self.btn_del_cat.clicked.connect(self._delete_category)
        cat_bar.addWidget(self.btn_new_cat)
        cat_bar.addWidget(self.btn_edit_cat)
        cat_bar.addWidget(self.btn_del_cat)
        root.addLayout(cat_bar)

        # Acciones de ítem + Import/Export
        actions = QHBoxLayout()
        self.btn_new_item = QPushButton("Nuevo Ítem")
        self.btn_edit_item = QPushButton("Editar Ítem")
        self.btn_del_item = QPushButton("Eliminar Ítem")
        self.btn_new_item.clicked.connect(self._new_item)
        self.btn_edit_item.clicked.connect(self._edit_item)
        self.btn_del_item.clicked.connect(self._delete_item)

        self.btn_import_excel = QPushButton("Importar Excel")
        self.btn_template_excel = QPushButton("Generar Plantilla Excel")
        self.btn_export_excel = QPushButton("Exportar a Excel")  # NUEVO

        self.btn_import_excel.clicked.connect(self._open_import_dialog)
        self.btn_template_excel.clicked.connect(self._generate_template)
        self.btn_export_excel.clicked.connect(self._export_to_excel)  # NUEVO

        actions.addWidget(self.btn_new_item)
        actions.addWidget(self.btn_edit_item)
        actions.addWidget(self.btn_del_item)
        actions.addWidget(self.btn_import_excel)
        actions.addWidget(self.btn_template_excel)
        actions.addWidget(self.btn_export_excel)
        root.addLayout(actions)

        # Búsqueda
        search_bar = QHBoxLayout()
        search_bar.addWidget(QLabel("Buscar:"))
        self.search_edit = QLineEdit(); self.search_edit.setPlaceholderText("Código, nombre o categoría…")
        self.search_edit.textChanged.connect(self._load_items)
        search_bar.addWidget(self.search_edit)
        root.addLayout(search_bar)

        # Tabla
        self.table = QTableWidget(0, 8)
        self.table.setHorizontalHeaderLabels(["#", "Código", "Nombre", "UD", "Costo", "Precio Venta", "Categoría", "Descripción"])
        header = self.table.horizontalHeader()
        header.setSectionResizeMode(QHeaderView.ResizeMode.Interactive)   # El usuario puede ajustar
        header.setSectionResizeMode(2, QHeaderView.ResizeMode.Stretch)    # Nombre se estira
        self.table.verticalHeader().setVisible(False)
        self.table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self.table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        # Ocultar columna Descripción (index 7) en la vista
        self.table.setColumnHidden(7, True)
        # Doble clic para editar
        self.table.itemDoubleClicked.connect(lambda *_: self._edit_item())
        # Menú contextual (clic derecho)
        self.table.setContextMenuPolicy(Qt.ContextMenuPolicy.CustomContextMenu)
        self.table.customContextMenuRequested.connect(self._on_table_context_menu)

        root.addWidget(self.table)

    # Categorías
    def _fetch_categories(self) -> List[Tuple[int, str, str, int, str]]:
        db = get_db_path()
        ensure_items_schema(db)
        with sqlite3.connect(db) as conn:
            cur = conn.execute("SELECT id, name, IFNULL(code_prefix,''), IFNULL(next_seq,1), IFNULL(description,'') FROM categories ORDER BY name")
            return cur.fetchall()

    def _load_categories(self):
        cats = self._fetch_categories()
        self.cat_combo.blockSignals(True)
        self.cat_combo.clear()
        for cid, name, prefix, next_seq, _ in cats:
            self.cat_combo.addItem(f"{name} ({prefix})", cid)
        self.cat_combo.blockSignals(False)
        self._update_next_code_preview()

    def _on_category_changed(self, _idx: int):
        self._update_next_code_preview()
        self._load_items()

    def _update_next_code_preview(self):
        cid = self.cat_combo.currentData()
        code = get_next_code(get_db_path(), cid) if cid else None
        self.next_code_preview.setText(f"Siguiente código: {code or '-'}")

    def _new_category(self):
        db = get_db_path()
        dlg = CategoryDialog(db, parent=self)
        if dlg.exec() == QDialog.DialogCode.Accepted:
            data = dlg.get_data()
            if not data:
                return
            name, prefix, seq, desc = data
            with sqlite3.connect(db) as conn:
                try:
                    conn.execute("INSERT INTO categories (name, code_prefix, next_seq, description) VALUES (?,?,?,?)",
                                 (name, prefix, seq, desc))
                    conn.commit()
                except sqlite3.IntegrityError as e:
                    QMessageBox.warning(self, "Error", f"No se pudo crear la categoría:\n{e}")
            self._load_categories()

    def _edit_category(self):
        db = get_db_path()
        cid = self.cat_combo.currentData()
        if not cid:
            QMessageBox.information(self, "Categoría", "Seleccione una categoría."); return
        with sqlite3.connect(db) as conn:
            row = conn.execute("SELECT id, name, code_prefix, next_seq, IFNULL(description,'') FROM categories WHERE id=?", (cid,)).fetchone()
        dlg = CategoryDialog(db, row, self)
        if dlg.exec() == QDialog.DialogCode.Accepted:
            data = dlg.get_data()
            if not data:
                return
            name, prefix, seq, desc = data
            with sqlite3.connect(db) as conn:
                try:
                    conn.execute("UPDATE categories SET name=?, code_prefix=?, next_seq=?, description=? WHERE id=?",
                                 (name, prefix, seq, desc, cid))
                    conn.commit()
                except sqlite3.IntegrityError as e:
                    QMessageBox.warning(self, "Error", f"No se pudo actualizar la categoría:\n{e}")
            self._load_categories()

    def _delete_category(self):
        db = get_db_path()
        cid = self.cat_combo.currentData()
        if not cid:
            QMessageBox.information(self, "Categoría", "Seleccione una categoría."); return
        with sqlite3.connect(db) as conn:
            count = conn.execute("SELECT COUNT(*) FROM items WHERE category_id=?", (cid,)).fetchone()[0]
        if count > 0:
            QMessageBox.warning(self, "Eliminar", "No puede eliminar una categoría con ítems asociados.")
            return
        ok = QMessageBox.question(self, "Eliminar", "¿Eliminar la categoría seleccionada?",
                                  QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No)
        if ok != QMessageBox.StandardButton.Yes:
            return
        with sqlite3.connect(db) as conn:
            conn.execute("DELETE FROM categories WHERE id=?", (cid,))
            conn.commit()
        self._load_categories()

    # Ítems
    def _new_item(self):
        db = get_db_path()
        cats = [(cid, name, prefix, next_seq) for cid, name, prefix, next_seq, _ in self._fetch_categories()]
        if not cats:
            QMessageBox.information(self, "Ítems", "Cree una categoría primero."); return
        dlg = ItemDialog(db, cats, parent=self)
        idx = dlg.cat_combo.findData(self.cat_combo.currentData())
        if idx >= 0:
            dlg.cat_combo.setCurrentIndex(idx)
        if dlg.exec() != QDialog.DialogCode.Accepted:
            return
        data = dlg.get_data()
        if not data:
            return
        cid = data["category_id"]
        manual = data["manual_code"]
        code = data["code"]
        if not manual:
            code = get_next_code(db, cid)
            if not code:
                QMessageBox.warning(self, "Código", "No fue posible generar el código."); return
        try:
            with sqlite3.connect(db) as conn:
                conn.execute("""
                    INSERT INTO items (code, name, unit, cost, price, category_id, description)
                    VALUES (?,?,?,?,?,?,?)
                """, (code, data["name"], data["unit"], data["cost"], data["price"], cid, data["description"]))
                conn.commit()
        except sqlite3.IntegrityError:
            QMessageBox.warning(self, "Código duplicado", f"El código {code} ya existe.")
            return
        if not manual:
            bump_category_seq(db, cid)
        self._load_items()
        self._update_next_code_preview()

    def _get_selected_item(self) -> Optional[dict]:
        row = self.table.currentRow()
        if row < 0:
            return None
        return {
            "code": self.table.item(row, 1).text(),
            "name": self.table.item(row, 2).text(),
            "unit": self.table.item(row, 3).text(),
            "cost": float(self.table.item(row, 4).text().replace(",", "")),
            "price": float(self.table.item(row, 5).text().replace(",", "")),
            "category_name": self.table.item(row, 6).text(),
            "description": self.table.item(row, 7).text() if not self.table.isColumnHidden(7) else "",
        }

    def _edit_item(self):
        db = get_db_path()
        selected = self._get_selected_item()
        if not selected:
            QMessageBox.information(self, "Ítems", "Seleccione un ítem de la lista."); return
        with sqlite3.connect(db) as conn:
            row = conn.execute("""
                SELECT i.id, i.code, i.name, i.unit, i.cost, i.price, i.category_id, IFNULL(i.description,'')
                FROM items i
                WHERE i.code = ?
            """, (selected["code"],)).fetchone()
        if not row:
            QMessageBox.warning(self, "Ítems", "No se encontró el ítem en la base de datos."); return
        item = {
            "id": row[0], "code": row[1], "name": row[2], "unit": row[3],
            "cost": row[4], "price": row[5], "category_id": row[6], "description": row[7]
        }
        cats = [(cid, name, prefix, next_seq) for cid, name, prefix, next_seq, _ in self._fetch_categories()]
        dlg = ItemDialog(db, cats, item, self)
        if dlg.exec() != QDialog.DialogCode.Accepted:
            return
        data = dlg.get_data()
        if not data:
            return
        manual = data["manual_code"]
        new_code = data["code"] if manual else item["code"]
        try:
            with sqlite3.connect(db) as conn:
                if new_code != item["code"]:
                    exists = conn.execute("SELECT 1 FROM items WHERE code=?", (new_code,)).fetchone()
                    if exists:
                        QMessageBox.warning(self, "Código duplicado", f"El código {new_code} ya existe.")
                        return
                conn.execute("""
                    UPDATE items
                    SET code=?, name=?, unit=?, cost=?, price=?, category_id=?, description=?
                    WHERE id=?
                """, (new_code, data["name"], data["unit"], data["cost"], data["price"], data["category_id"], data["description"], item["id"]))
                conn.commit()
        except sqlite3.IntegrityError as e:
            QMessageBox.warning(self, "Error", f"No se pudo actualizar el ítem:\n{e}")
            return
        self._load_items()

    def _delete_item(self):
        db = get_db_path()
        selected = self._get_selected_item()
        if not selected:
            QMessageBox.information(self, "Ítems", "Seleccione un ítem de la lista."); return
        ok = QMessageBox.question(self, "Eliminar", f"¿Eliminar el ítem {selected['code']}?",
                                  QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No)
        if ok != QMessageBox.StandardButton.Yes:
            return
        with sqlite3.connect(db) as conn:
            conn.execute("DELETE FROM items WHERE code=?", (selected["code"],))
            conn.commit()
        self._load_items()
        self._update_next_code_preview()

    # Carga de tabla
    def _load_items(self):
        db = get_db_path()
        ensure_items_schema(db)
        search = (self.search_edit.text() or "").strip().lower()
        current_cid = self.cat_combo.currentData()
        query = """
            SELECT i.id, i.code, i.name, i.unit, i.cost, i.price, IFNULL(c.name,''), IFNULL(i.description,'')
            FROM items i
            LEFT JOIN categories c ON i.category_id = c.id
            WHERE (LOWER(i.code) LIKE ? OR LOWER(i.name) LIKE ? OR LOWER(IFNULL(c.name,'')) LIKE ?)
        """
        params = [f"%{search}%", f"%{search}%", f"%{search}%"]
        if current_cid:
            query += " AND i.category_id = ?"
            params.append(current_cid)
        query += " ORDER BY i.id"
        with sqlite3.connect(db) as conn:
            items = conn.execute(query, params).fetchall()

        self.table.setRowCount(0)
        for idx, row in enumerate(items, 1):
            r = self.table.rowCount()
            self.table.insertRow(r)
            # Columna # (4 dígitos máx. aprox.)
            self.table.setItem(r, 0, QTableWidgetItem(str(idx)))
            # code, name, unit, cost, price, category, description
            code, name, unit, cost, price, cat_name, desc = row[1], row[2], row[3], row[4], row[5], row[6], row[7]
            self.table.setItem(r, 1, QTableWidgetItem(str(code or "")))
            self.table.setItem(r, 2, QTableWidgetItem(str(name or "")))
            self.table.setItem(r, 3, QTableWidgetItem(str(unit or "")))
            # Formato 00,000.00 en costo y precio
            try:
                cost_fmt = f"{float(cost or 0):,.2f}"
            except Exception:
                cost_fmt = "0.00"
            try:
                price_fmt = f"{float(price or 0):,.2f}"
            except Exception:
                price_fmt = "0.00"
            self.table.setItem(r, 4, QTableWidgetItem(cost_fmt))
            self.table.setItem(r, 5, QTableWidgetItem(price_fmt))
            self.table.setItem(r, 6, QTableWidgetItem(str(cat_name or "")))
            self.table.setItem(r, 7, QTableWidgetItem(str(desc or "")))

        # Ocultar descripción en vista (ya se configuró, reafirmar)
        self.table.setColumnHidden(7, True)
        # Aplicar layout de columnas (anchos)
        self._apply_table_column_layout()

    # --- Layout de columnas conforme requerimientos ---
    def _apply_table_column_layout(self):
        """
        Ajusta anchos:
        - #: ~4 dígitos
        - Código: ~7 dígitos
        - Nombre: se estira (stretch)
        - UD: ~4 chars
        - Costo/Precio: ~9-12 dígitos con separadores
        - Categoría: dejar por defecto
        - Descripción: oculta
        """
        fm = QFontMetrics(self.table.font())
        hash_w = fm.horizontalAdvance("9999") + 20
        code_w = fm.horizontalAdvance("9999999") + 22
        ud_w   = fm.horizontalAdvance("UNID") + 22
        money_w = fm.horizontalAdvance("9,999,999.99") + 28

        header = self.table.horizontalHeader()
        header.resizeSection(0, max(60, hash_w))     # #
        header.resizeSection(1, max(90, code_w))     # Código
        # 2 = Nombre (stretch)
        header.resizeSection(3, max(70, ud_w))       # UD
        header.resizeSection(4, max(110, money_w))   # Costo
        header.resizeSection(5, max(120, money_w))   # Precio Venta
        # 6 = Categoría
        # 7 = Descripción (oculta)
        self.table.setColumnHidden(7, True)

    # --- Menú contextual tabla ---
    def _on_table_context_menu(self, pos):
        index = self.table.indexAt(pos)
        if not index.isValid():
            return
        # Seleccionar fila
        self.table.selectRow(index.row())
        self.table.setCurrentCell(index.row(), 0)

        menu = QMenu(self)
        act_view = QAction("Visualizar", self)
        act_edit = QAction("Editar", self)
        act_del  = QAction("Eliminar", self)

        act_view.triggered.connect(self._view_item)
        act_edit.triggered.connect(self._edit_item)
        act_del.triggered.connect(self._delete_item)

        menu.addAction(act_view)
        menu.addSeparator()
        menu.addAction(act_edit)
        menu.addAction(act_del)

        menu.exec(self.table.viewport().mapToGlobal(pos))

    def _view_item(self):
        data = self._get_selected_item()
        if not data:
            QMessageBox.information(self, "Ítems", "Seleccione un ítem."); return
        # Mostrar un resumen en un cuadro de diálogo
        text = (
            f"Código: {data['code']}\n"
            f"Nombre: {data['name']}\n"
            f"Unidad: {data['unit']}\n"
            f"Categoría: {data['category_name']}\n"
            f"Costo: {float(data['cost']):,.2f}\n"
            f"Precio Venta: {float(data['price']):,.2f}\n"
            f"Descripción: {(data.get('description') or '')}\n"
        )
        QMessageBox.information(self, "Ítem", text)

    # Importar / Exportar / Plantilla
    def _open_import_dialog(self):
        dlg = ExcelImportDialog(self)
        dlg.exec()
        # tras importar, refrescar
        self._load_categories()
        self._load_items()

    def _generate_template(self):
        if Workbook is None:
            QMessageBox.critical(self, "Dependencia faltante", "Instala openpyxl: pip install openpyxl")
        fn, _ = QFileDialog.getSaveFileName(self, "Guardar plantilla Excel", "plantilla_items.xlsx", "Excel (*.xlsx)")
        if not fn:
            return
        try:
            generate_excel_template(fn)
            QMessageBox.information(self, "Plantilla", f"Plantilla creada en:\n{fn}")
        except Exception as e:
            QMessageBox.critical(self, "Plantilla", f"No se pudo generar la plantilla:\n{e}")

    def _export_to_excel(self):
        if Workbook is None:
            QMessageBox.critical(self, "Dependencia faltante", "Instala openpyxl: pip install openpyxl")
            return
        fn, _ = QFileDialog.getSaveFileName(self, "Exportar a Excel", "inventario_items.xlsx", "Excel (*.xlsx)")
        if not fn:
            return
        # Barra de progreso (simple: 3 pasos)
        progress = QProgressDialog("Exportando a Excel…", None, 0, 3, self)
        progress.setWindowModality(Qt.WindowModality.ApplicationModal)
        progress.setMinimumDuration(200)
        try:
            progress.setValue(1)
            export_current_to_excel(fn)
            progress.setValue(3)
            QMessageBox.information(self, "Exportación", f"Archivo Excel exportado en:\n{fn}")
        except Exception as e:
            QMessageBox.critical(self, "Exportación", f"No se pudo exportar:\n{e}")
        finally:
            progress.setValue(3)