from __future__ import annotations

"""
utils/quotation_templates.py

Plantillas y generadores para Cotizaciones (Excel + PDF).

Incluye:
- generate_quotation_template(path)
- generate_quotation_excel(quotation_data, items, save_path, company_name="", template=None)
- generate_quotation_pdf(quotation_data, items, save_path, company_name="", template=None)

Notas:
- Soporta inyección de "template" (dict) con keys: primary_color, secondary_color, show_logo, logo_path, footer_lines, itbis_rate, etc.
- Si template es None y quotation_data contiene company_id, intenta cargar la plantilla con load_template(company_id).
- Requiere openpyxl para Excel y reportlab (más Pillow opcional) para PDF.
"""

import os
import datetime
from typing import List, Dict, Optional

# Escape para reportlab (sanitizar texto en Paragraph HTML)
from xml.sax.saxutils import escape

# --------------------
# openpyxl (Excel)
# --------------------
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter
    from openpyxl.drawing.image import Image as XLImage
except Exception:
    Workbook = None
    XLImage = None

# --------------------
# Pillow (opcional)
# --------------------
try:
    from PIL import Image as PILImage
except Exception:
    PILImage = None

# --------------------
# reportlab (PDF)
# --------------------
REPORTLAB_AVAILABLE = False
try:
    from reportlab.lib.pagesizes import letter
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import (
        SimpleDocTemplate,
        Table,
        TableStyle,
        Paragraph,
        Spacer,
        Image as RLImage,
    )
    from reportlab.lib.units import mm
    from reportlab.lib.enums import TA_LEFT, TA_RIGHT
    REPORTLAB_AVAILABLE = True
except Exception:
    # reportlab is optional; the module still works for Excel-only setups
    REPORTLAB_AVAILABLE = False

# --------------------
# Template manager (optional, used to load saved templates)
# --------------------
try:
    from utils.template_manager import get_data_root, load_template
except Exception:
    # If template_manager is not available, provide fallback stubs
    def get_data_root():
        return os.getcwd()

    def load_template(company_id: int):
        return {}

# --------------------
# Helpers
# --------------------
def _resolve_template(data: Dict, template: Optional[Dict]) -> Dict:
    """Return a template dict with defaults merged with given template or loaded one."""
    tpl = template
    if tpl is None and data.get("company_id") is not None:
        try:
            tpl = load_template(int(data.get("company_id")))
        except Exception:
            tpl = {}
    if tpl is None:
        tpl = {}
    defaults = {
        "primary_color": "#1f7a44",  # verde por defecto
        "secondary_color": "#EEF8F1",
        "show_logo": True,
        "logo_path": "",
        "header_lines": [],
        "footer_lines": [],
        "font_name": "Helvetica",
        "font_size": 10,
        "logo_width_px": 160,
        "itbis_rate": 0.18,
    }
    out = dict(defaults)
    out.update(tpl or {})
    return out


def _resolve_logo_path_rel_to_abs(logo_rel: str) -> Optional[str]:
    """Resolve a possibly-relative logo path to an absolute filesystem path, if exists."""
    if not logo_rel:
        return None
    # Accept file:/// prefix
    if logo_rel.startswith("file:///"):
        candidate = logo_rel.replace("file:///", "")
        if os.path.exists(candidate):
            return candidate
    if os.path.isabs(logo_rel) and os.path.exists(logo_rel):
        return logo_rel
    # Try relative to data root if available
    try:
        data_root = get_data_root()
        abs_path = os.path.join(data_root, logo_rel)
        if os.path.exists(abs_path):
            return abs_path
    except Exception:
        pass
    # Finally try relative to cwd
    if os.path.exists(logo_rel):
        return os.path.abspath(logo_rel)
    return None


def _hex_to_reportlab_color(hex_color: str):
    try:
        return colors.HexColor(hex_color)
    except Exception:
        return colors.HexColor("#1f7a44")


def _fmt_currency(n):
    try:
        return "{:,.2f}".format(float(n or 0))
    except Exception:
        return "0.00"


# ------------------------------------------------------------------
# Plantilla Excel (generador simple)
# ------------------------------------------------------------------
def generate_quotation_template(path: str):
    """
    Genera una plantilla Excel básica para cotizaciones.
    """
    if Workbook is None:
        raise RuntimeError("openpyxl no está instalado. Instala con: pip install openpyxl")

    wb = Workbook()
    sh = wb.active
    sh.title = "Cotizacion"

    # Encabezado de la plantilla (campos a completar)
    sh["A1"] = "company_name"
    sh["B1"] = "quotation_date"
    sh["C1"] = "client_name"
    sh["D1"] = "client_rnc"
    sh["E1"] = "currency"
    sh["F1"] = "notes"

    # Tabla de items (cabeceras)
    start = 3
    sh.cell(row=start, column=1, value="code (optional)")
    sh.cell(row=start, column=2, value="description")
    sh.cell(row=start, column=3, value="unit")
    sh.cell(row=start, column=4, value="quantity")
    sh.cell(row=start, column=5, value="unit_price")
    sh.cell(row=start, column=6, value="subtotal (optional)")

    # Ejemplo
    sh.cell(row=start + 1, column=2, value="Servicio técnico por hora")
    sh.cell(row=start + 1, column=3, value="HR")
    sh.cell(row=start + 1, column=4, value=1)
    sh.cell(row=start + 1, column=5, value=1500)

    notes = wb.create_sheet("Notas")
    notes["A1"] = "Instrucciones"
    notes["A2"] = "- Rellena la fila 1: company_name, quotation_date (YYYY-MM-DD), client_name, client_rnc, currency, notes."
    notes["A3"] = "- En 'Cotizacion' debajo de la fila de cabecera agrega las filas de items."

    for col, width in enumerate([14, 50, 12, 12, 18, 18], start=1):
        try:
            sh.column_dimensions[get_column_letter(col)].width = width
        except Exception:
            pass

    wb.save(path)


# ------------------------------------------------------------------
# Generador Excel (mantener compatibilidad)
# ------------------------------------------------------------------
def generate_quotation_excel(
    quotation_data: Dict,
    items: List[Dict],
    save_path: str,
    company_name: str = "",
    template: Dict = None,
):
    """
    Genera un Excel formateado de la Cotización.
    - Usa openpyxl.
    - Inserta logo (si existe) sin fallar por temporales en Windows.
      Mantiene los archivos temporales hasta después de wb.save(...).
    - Cantidad SIEMPRE con 2 decimales.
    - Totales alineados a la derecha.
    - Colores/estilos tomados del template cuando existan.
    """
    if Workbook is None:
        raise RuntimeError("openpyxl no está instalado. Instala con: pip install openpyxl")

    # -------------------------
    # Helpers locales (auto-contenidos)
    # -------------------------
    import re, tempfile, shutil, urllib.parse

    def _file_uri_to_path(uri_or_path: str) -> str:
        s = (uri_or_path or "").strip()
        if not s:
            return ""
        if s.lower().startswith("file:///"):
            return urllib.parse.unquote(s.replace("file:///", ""))
        return s

    def _abs_path(path: str) -> str:
        p = _file_uri_to_path(path)
        if not p:
            return ""
        # Intenta resolver relativo a data root si está disponible
        try:
            from utils.template_manager import get_data_root  # opcional
            root = get_data_root()
            cand = os.path.join(root, p)
            if os.path.exists(cand):
                return os.path.abspath(cand)
        except Exception:
            pass
        if os.path.isabs(p):
            return p
        # Relativo al cwd
        cand = os.path.join(os.getcwd(), p)
        return os.path.abspath(cand)

    def _persist_temp_copy(src_path: str, suffix: str = ".png") -> str:
        """
        Copia el archivo a un temporal PERSISTENTE (delete=False) y cierra el handle,
        para que openpyxl pueda leerlo al guardar el .xlsx en Windows.
        """
        if not src_path or not os.path.exists(src_path):
            return ""
        fd, dst = tempfile.mkstemp(prefix="facot_", suffix=suffix, dir=tempfile.gettempdir())
        os.close(fd)  # imprescindible en Windows
        shutil.copy2(src_path, dst)
        return dst

    # -------------------------
    # Preparar template y libro
    # -------------------------
    template = _resolve_template(quotation_data or {}, template)

    wb = Workbook()
    sh = wb.active
    sh.title = "Cotizacion"

    bold = Font(bold=True)
    primary_color = template.get("primary_color", "#1f7a44")
    header_fill = PatternFill("solid", fgColor=primary_color.replace("#", ""))
    border = Border(left=Side(style="thin"), right=Side(style="thin"),
                    top=Side(style="thin"), bottom=Side(style="thin"))

    temp_files_to_cleanup: List[str] = []

    # -------------------------
    # Logo (robusto en Windows)
    # -------------------------
    try:
        want_logo = bool(template.get("show_logo"))
        raw_logo = template.get("logo_path") or quotation_data.get("logo_path") or ""
        if want_logo and raw_logo and XLImage is not None:
            # Resolver a ruta absoluta local (acepta file:/// y relativas)
            logo_abs = _abs_path(raw_logo)
            if logo_abs and os.path.exists(logo_abs):
                logo_w = int(template.get("logo_width_px", 160) or 160)
                try:
                    if PILImage is not None:
                        pil = PILImage.open(logo_abs)
                        ratio = (pil.height / pil.width) if pil.width else 1.0
                        new_h = max(1, int(logo_w * ratio))
                        # Crear temporal persistente, guardar y cerrar antes de usar en openpyxl
                        tmp_resized = _persist_temp_copy(logo_abs, ".png")
                        temp_files_to_cleanup.append(tmp_resized)
                        try:
                            pil_resized = pil.resize((logo_w, new_h))
                            pil_resized.save(tmp_resized, format="PNG")
                        except Exception:
                            # Si resize falla, dejamos la copia original ya creada
                            pass
                        img = XLImage(tmp_resized)
                        sh.add_image(img, "A1")
                    else:
                        # Sin PIL: insertar original (o su copia) tal cual
                        tmp_copy = _persist_temp_copy(logo_abs, os.path.splitext(logo_abs)[1] or ".png")
                        if tmp_copy:
                            temp_files_to_cleanup.append(tmp_copy)
                            img = XLImage(tmp_copy)
                            sh.add_image(img, "A1")
                except Exception:
                    # Ignorar errores de imagen
                    pass
    except Exception:
        pass

    # -------------------------
    # Encabezado textual
    # -------------------------
    sh.merge_cells("A1:F1")
    sh["A1"] = company_name or quotation_data.get("company_name", "")
    sh["A1"].font = Font(size=14, bold=True)

    sh["A2"] = "Cotización:"
    sh["B2"] = quotation_data.get("quotation_date", datetime.date.today().isoformat())
    sh["A3"] = "Cliente:"
    sh["B3"] = quotation_data.get("client_name", "")
    sh["A4"] = "RNC/Cédula:"
    sh["B4"] = quotation_data.get("client_rnc", "")
    sh["A5"] = "Moneda:"
    sh["B5"] = quotation_data.get("currency", "RD$")

    # -------------------------
    # Tabla de ítems
    # -------------------------
    start_row = 8
    headers = ["#", "Código", "Descripción", "Unidad", "Cantidad", "Precio Unitario", "Subtotal"]
    for idx, h in enumerate(headers, start=1):
        c = sh.cell(row=start_row, column=idx, value=h)
        c.font = bold
        try:
            if primary_color:
                c.fill = PatternFill("solid", fgColor=primary_color.replace("#", ""))
        except Exception:
            c.fill = header_fill
        c.alignment = Alignment(horizontal="center")
        c.border = border

    row = start_row + 1
    total = 0.0
    for i, it in enumerate(items, start=1):
        code = it.get("code", "") or it.get("item_code", "")
        desc = it.get("description", "") or it.get("nombre", "")
        unit = it.get("unit", "") or it.get("unidad", "")
        try:
            qty = float(it.get("quantity", it.get("cantidad", 0)) or 0)
        except Exception:
            qty = 0.0
        try:
            price = float(it.get("unit_price", it.get("precio", 0)) or 0)
        except Exception:
            price = 0.0
        subtotal = qty * price
        total += subtotal

        sh.cell(row=row, column=1, value=i).border = border
        sh.cell(row=row, column=2, value=code).border = border
        sh.cell(row=row, column=3, value=desc).border = border
        sh.cell(row=row, column=4, value=unit).border = border

        qcell = sh.cell(row=row, column=5, value=qty)
        qcell.number_format = '#,##0.00'          # SIEMPRE 2 decimales
        qcell.alignment = Alignment(horizontal="right")
        qcell.border = border

        pcell = sh.cell(row=row, column=6, value=price)
        pcell.number_format = '#,##0.00'
        pcell.alignment = Alignment(horizontal="right")
        pcell.border = border

        scell = sh.cell(row=row, column=7, value=subtotal)
        scell.number_format = '#,##0.00'
        scell.alignment = Alignment(horizontal="right")
        scell.border = border

        row += 1

    # -------------------------
    # Totales (alineados a la derecha)
    # -------------------------
    itbis_rate = float(quotation_data.get("itbis_rate", template.get("itbis_rate", 0.0)) or 0.0)
    apply_itbis = quotation_data.get("apply_itbis")
    if apply_itbis is None:
        apply_itbis = itbis_rate > 0

    # Etiquetas columna F, valores columna G
    lbl_sub = sh.cell(row=row + 1, column=6, value="Subtotal:")
    lbl_sub.font = Font(bold=True)
    lbl_sub.alignment = Alignment(horizontal="right")
    val_sub = sh.cell(row=row + 1, column=7, value=total)
    val_sub.number_format = '#,##0.00'
    val_sub.alignment = Alignment(horizontal="right")

    itbis = total * itbis_rate if apply_itbis else 0.0
    if itbis > 0:
        lbl_itb = sh.cell(row=row + 2, column=6, value=f"ITBIS ({itbis_rate * 100:.0f}%):")
        lbl_itb.font = Font(bold=True)
        lbl_itb.alignment = Alignment(horizontal="right")
        val_itb = sh.cell(row=row + 2, column=7, value=itbis)
        val_itb.number_format = '#,##0.00'
        val_itb.alignment = Alignment(horizontal="right")

        lbl_tot = sh.cell(row=row + 3, column=6, value="Total:")
        lbl_tot.font = Font(bold=True)
        lbl_tot.alignment = Alignment(horizontal="right")
        val_tot = sh.cell(row=row + 3, column=7, value=total + itbis)
        val_tot.number_format = '#,##0.00'
        val_tot.alignment = Alignment(horizontal="right")
    else:
        lbl_tot = sh.cell(row=row + 2, column=6, value="Total:")
        lbl_tot.font = Font(bold=True)
        lbl_tot.alignment = Alignment(horizontal="right")
        val_tot = sh.cell(row=row + 2, column=7, value=total)
        val_tot.number_format = '#,##0.00'
        val_tot.alignment = Alignment(horizontal="right")

    # -------------------------
    # Notas
    # -------------------------
    notes = quotation_data.get("notes", "")
    if notes:
        nr = row + 5
        sh.cell(row=nr, column=1, value="Notas:")
        sh.cell(row=nr + 1, column=1, value=notes)

    # -------------------------
    # Anchos de columna
    # -------------------------
    widths = [6, 14, 50, 10, 12, 18, 18]
    for i, w in enumerate(widths, start=1):
        try:
            sh.column_dimensions[get_column_letter(i)].width = w
        except Exception:
            pass

    # -------------------------
    # Guardar y limpiar temporales
    # -------------------------
    try:
        wb.save(save_path)
    finally:
        for tmp in temp_files_to_cleanup:
            try:
                if tmp and os.path.exists(tmp):
                    os.remove(tmp)
            except Exception:
                pass

import os, tempfile, shutil, urllib.parse

def file_uri_to_path(uri_or_path: str) -> str:
    """
    Convierte file:///C:/... a C:/..., deja absoluto tal cual, y si es relativo lo retorna relativo.
    """
    s = (uri_or_path or "").strip()
    if not s:
        return ""
    if s.lower().startswith("file:///"):
        # file:///C:/... -> C:/...
        return urllib.parse.unquote(s.replace("file:///", ""))
    return s

def persist_temp_copy(src_path: str, suffix: str = ".png") -> str:
    """
    Crea una copia temporal PERSISTENTE del archivo (delete=False).
    Retorna la ruta al archivo copiado dentro del directorio temporal del sistema.
    """
    if not src_path or not os.path.exists(src_path):
        return ""
    tmp_dir = tempfile.gettempdir()
    fd, dst = tempfile.mkstemp(prefix="facot_", suffix=suffix, dir=tmp_dir)
    os.close(fd)  # cerramos el handle para poder reabrirlo con openpyxl/xlsxwriter
    shutil.copy2(src_path, dst)
    return dst


# ------------------------------------------------------------------
# Generador PDF (reportlab)
# ------------------------------------------------------------------
def generate_quotation_pdf(
    quotation_data: Dict,
    items: List[Dict],
    save_path: str,
    company_name: str = "",
    template: Optional[Dict] = None,
):
    """
    Genera un PDF de cotización usando reportlab ajustado al layout HTML:
    - tamaño Letter
    - color primario tomado de template['primary_color'] o company.primary_color
    - cabecera con logo y datos de empresa en MAYÚSCULAS
    - tabla de items con encabezado coloreado
    - totales alineados a la derecha
    - dos firmas al pie (Firma Autorizada / Recibido Por)
    """
    if not REPORTLAB_AVAILABLE:
        raise RuntimeError("reportlab no está instalado. Instala con: pip install reportlab")

    tpl = _resolve_template(quotation_data or {}, template)

    primary_hex = tpl.get("primary_color") or "#1f7a44"
    secondary_hex = tpl.get("secondary_color") or "#EEF8F1"
    primary_color = _hex_to_reportlab_color(primary_hex)
    secondary_color = _hex_to_reportlab_color(secondary_hex)

    # Documento
    doc = SimpleDocTemplate(
        save_path,
        pagesize=letter,
        leftMargin=15 * mm,
        rightMargin=15 * mm,
        topMargin=15 * mm,
        bottomMargin=15 * mm,
    )

    styles = getSampleStyleSheet()
    normal = styles["Normal"]
    normal.fontName = "Helvetica"
    normal.fontSize = 10
    bold_style = ParagraphStyle("bold", parent=normal, fontName="Helvetica-Bold")
    small_muted = ParagraphStyle("muted_small", parent=normal, textColor=colors.HexColor("#6b7280"), fontSize=9)
    company_style = ParagraphStyle("company", parent=bold_style, fontSize=12, alignment=TA_LEFT)
    meta_style = ParagraphStyle("meta", parent=normal, fontSize=10, textColor=colors.HexColor("#6b7280"))
    right_bold = ParagraphStyle("right_bold", parent=bold_style, alignment=TA_RIGHT)
    total_label_style = ParagraphStyle("total_label", parent=bold_style, alignment=TA_RIGHT, fontSize=10)
    total_value_style = ParagraphStyle("total_value", parent=right_bold, textColor=primary_color, fontSize=12)

    story = []

    # Company data shape support
    company = {}
    raw_company = quotation_data.get("company")
    if isinstance(raw_company, dict) and raw_company:
        company = raw_company
    else:
        # fallback to top-level fields
        company = {
            "name": quotation_data.get("company_name") or company_name or "",
            "rnc": quotation_data.get("company_rnc") or "",
            "address_line1": quotation_data.get("company_address") or quotation_data.get("company_address_line1") or "",
            "address_line2": quotation_data.get("company_address_line2") or "",
            "phone": quotation_data.get("company_phone") or quotation_data.get("company_telefono") or "",
            "logo_path": quotation_data.get("company_logo") or "",
        }

    # Resolve logo
    logo_flowable = None
    logo_path = (company.get("logo_path") or tpl.get("logo_path") or "")
    logo_fs = None
    if logo_path:
        if logo_path.startswith("file:///"):
            logo_fs = logo_path.replace("file:///", "")
        else:
            logo_fs = logo_path
        if not os.path.exists(logo_fs):
            # try resolve relative
            logo_fs = _resolve_logo_path_rel_to_abs(logo_path)
    if logo_fs and os.path.exists(logo_fs):
        try:
            # RLImage expects path and width/height in points or 'kind' param
            logo_flowable = RLImage(logo_fs, width=120, height=68)
        except Exception:
            logo_flowable = None

    # Header table: left column company, right column meta
    left_parts = []
    if logo_flowable:
        left_parts.append(logo_flowable)

    comp_lines = []
    comp_name = (company.get("name") or "").upper()
    if comp_name:
        comp_lines.append(Paragraph(comp_name, company_style))
    if company.get("rnc"):
        comp_lines.append(Paragraph(f"RNC: {escape(str(company.get('rnc')))}", meta_style))
    if company.get("address_line1"):
        comp_lines.append(Paragraph(escape(str(company.get("address_line1"))), small_muted))
    if company.get("address_line2"):
        comp_lines.append(Paragraph(escape(str(company.get("address_line2"))), small_muted))
    if company.get("phone"):
        comp_lines.append(Paragraph(escape(str(company.get("phone"))), small_muted))

    left_cell = left_parts + comp_lines if left_parts else comp_lines

    qnum = quotation_data.get("number") or quotation_data.get("quotation_number") or quotation_data.get("id") or ""
    qdate = quotation_data.get("date") or quotation_data.get("quotation_date") or ""
    right_cell = [
        Paragraph(f'<font color="{primary_hex}"><b>COTIZACIÓN N°: {escape(str(qnum))}</b></font>', normal),
        Paragraph(f'FECHA: {escape(str(qdate))}', meta_style),
    ]

    header_table = Table([[left_cell, right_cell]], colWidths=[doc.width * 0.65, doc.width * 0.35])
    header_table.setStyle(
        TableStyle(
            [
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("LEFTPADDING", (0, 0), (-1, -1), 0),
                ("RIGHTPADDING", (0, 0), (-1, -1), 0),
            ]
        )
    )
    story.append(header_table)
    story.append(Spacer(1, 8))

    # Client block with background color (approximate using a Table cell background)
    client_name = quotation_data.get("client_name") or quotation_data.get("customer_name") or ""
    client_rnc = quotation_data.get("client_rnc") or quotation_data.get("customer_rnc") or ""
    client_block = [
        Paragraph('<b>DATOS DEL CLIENTE</b>', bold_style),
        Spacer(1, 4),
        Paragraph(f'<b>NOMBRE O RAZÓN SOCIAL:</b> {escape(str(client_name))}', normal),
        Paragraph(f'<b>RNC CLIENTE:</b> {escape(str(client_rnc))}', normal),
    ]
    client_table = Table([[client_block]], colWidths=[doc.width])
    client_table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, -1), secondary_color),
                ("LEFTPADDING", (0, 0), (-1, -1), 8),
                ("RIGHTPADDING", (0, 0), (-1, -1), 8),
                ("TOPPADDING", (0, 0), (-1, -1), 8),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
            ]
        )
    )
    story.append(client_table)
    story.append(Spacer(1, 10))

    # Items table
    data = []
    # header row
    data.append(
        [
            Paragraph("Cant.", bold_style),
            Paragraph("Código / Descripción", bold_style),
            Paragraph("Unidad", bold_style),
            Paragraph("Valor (RD$)", bold_style),
            Paragraph("Subtotal (RD$)", bold_style),
        ]
    )

    for it in items or []:
        code = it.get("code") or it.get("codigo") or ""
        desc = it.get("description") or it.get("descripcion") or ""
        qty = it.get("quantity") or it.get("cantidad") or 0
        unit = it.get("unit") or it.get("unidad") or ""
        up = it.get("unit_price") or it.get("unitario") or 0
        try:
            qtyf = float(qty or 0)
        except Exception:
            qtyf = 0.0
        try:
            upf = float(up or 0)
        except Exception:
            upf = 0.0
        line = qtyf * upf

        data.append(
            [
                Paragraph(str(qtyf), normal),
                Paragraph(f"<b>{escape(str(code))}</b><br/><font color='#6b7280'>{escape(str(desc))}</font>", normal),
                Paragraph(str(unit), normal),
                Paragraph(_fmt_currency(upf), normal),
                Paragraph(_fmt_currency(line), normal),
            ]
        )

    table = Table(
        data,
        colWidths=[doc.width * 0.08, doc.width * 0.52, doc.width * 0.10, doc.width * 0.15, doc.width * 0.15],
    )
    table_style = TableStyle(
        [
            ("GRID", (0, 0), (-1, -1), 0, colors.white),
            ("LINEBELOW", (0, 0), (-1, 0), 0.5, primary_color),
            ("BACKGROUND", (0, 0), (-1, 0), primary_color),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("ALIGN", (0, 0), (0, -1), "CENTER"),
            ("ALIGN", (3, 1), (4, -1), "RIGHT"),
            ("LEFTPADDING", (0, 0), (-1, -1), 6),
            ("RIGHTPADDING", (0, 0), (-1, -1), 6),
        ]
    )
    table.setStyle(table_style)
    story.append(table)
    story.append(Spacer(1, 8))

    # Totals
    subtotal = sum(
        (float(it.get("quantity") or it.get("cantidad") or 0) * float(it.get("unit_price") or it.get("unitario") or 0))
        for it in items or []
    )
    itbis_rate = float(tpl.get("itbis_rate") or quotation_data.get("itbis_rate") or 0.18)
    itbis_amount = subtotal * itbis_rate
    total = subtotal + itbis_amount

    totals_data = [
        [Paragraph("SUBTOTAL RD$:", normal), Paragraph(_fmt_currency(subtotal), right_bold)],
        [Paragraph(f"ITBIS ({int(itbis_rate * 100)}%):", normal), Paragraph(_fmt_currency(itbis_amount), right_bold)],
        [Paragraph("<b>TOTAL RD$:</b>", total_label_style), Paragraph(_fmt_currency(total), total_value_style)],
    ]
    totals_table = Table(totals_data, colWidths=[doc.width * 0.6, doc.width * 0.4])
    totals_table.setStyle(
        TableStyle(
            [
                ("ALIGN", (1, 0), (1, -1), "RIGHT"),
                ("LEFTPADDING", (0, 0), (-1, -1), 6),
                ("RIGHTPADDING", (0, 0), (-1, -1), 6),
                ("TOPPADDING", (0, 0), (-1, -1), 2),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
                ("LINEABOVE", (0, 2), (-1, 2), 0.5, colors.lightgrey),
            ]
        )
    )
    story.append(Spacer(1, 6))
    story.append(totals_table)

    # Footer / conditions
    footer_lines = tpl.get("footer_lines") or []
    if footer_lines:
        story.append(Spacer(1, 8))
        for line in footer_lines:
            story.append(Paragraph(line, small_muted))

    # Spacer to push signatures to bottom (heuristic)
    story.append(Spacer(1, 80 * mm))

    # Signature blocks
    sig_table = Table(
        [
            [
                Paragraph("<para align='center'><u>&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;</u></para>", normal),
                Paragraph("<para align='center'><u>&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;&nbsp;</u></para>", normal),
            ],
            [Paragraph("<para align='center'><b>Firma Autorizada</b></para>", normal), Paragraph("<para align='center'><b>Recibido Por</b></para>", normal)],
        ],
        colWidths=[doc.width * 0.45, doc.width * 0.45],
    )
    sig_table.setStyle(
        TableStyle(
            [
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("TOPPADDING", (0, 0), (-1, -1), 6),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
            ]
        )
    )
    story.append(sig_table)

    # Build PDF
    doc.build(story)
    return True