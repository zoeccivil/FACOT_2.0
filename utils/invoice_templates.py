
from __future__ import annotations

import os
import datetime
from typing import List, Dict, Optional

# Excel
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, NamedStyle
    from openpyxl.utils import get_column_letter
    from openpyxl.drawing.image import Image as XLImage
except Exception:
    Workbook = None
    XLImage = None

# Pillow (opcional, para redimensionar logos)
try:
    from PIL import Image as PILImage
except Exception:
    PILImage = None

# PDF (optional)
try:
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image as RLImage
    from reportlab.lib.utils import ImageReader
    REPORTLAB_AVAILABLE = True
except Exception:
    REPORTLAB_AVAILABLE = False

# helper to resolve template data root
from utils.template_manager import get_data_root, load_template

def _resolve_template(data: Dict, template: Optional[Dict]) -> Dict:
    tpl = template
    if tpl is None and data.get("company_id") is not None:
        tpl = load_template(int(data.get("company_id")))
    if tpl is None:
        tpl = {}
    # return merged with minimal defaults to avoid KeyError
    defaults = {
        "show_logo": True,
        "logo_path": "",
        "header_lines": ["", "", ""],
        "footer_lines": ["", ""],
        "primary_color": "#1f4e79",
        "font_name": "Arial",
        "font_size": 10,
        "logo_width_px": 180,
        "layout": "default"
    }
    out = dict(defaults)
    out.update(tpl)
    return out

def _resolve_logo_path_rel_to_abs(logo_rel: str) -> Optional[str]:
    if not logo_rel:
        return None
    # if it's already absolute path, prefer it
    if os.path.isabs(logo_rel) and os.path.exists(logo_rel):
        return logo_rel
    # else resolve relative to data root
    data_root = get_data_root()
    abs_path = os.path.join(data_root, logo_rel)
    if os.path.exists(abs_path):
        return abs_path
    # try fallback: maybe user provided path relative to cwd
    if os.path.exists(logo_rel):
        return os.path.abspath(logo_rel)
    return None

def generate_invoice_excel(invoice_data: Dict, items: List[Dict], save_path: str, company_name: str = "", template: Dict = None):
    """
    Igual que antes; este recordatorio es solo para que confirmes que
    invoice_data puede traer 'due_date' y lo mostrará en el encabezado si lo agregas.
    Si ya lo tienes en el PDF, no hace falta imprimirlo en Excel, pero puedes añadir:
        sh["C3"] = "Vencimiento:"
        sh["D3"] = invoice_data.get("due_date") or ""
    """
    ...
    if Workbook is None:
        raise RuntimeError("openpyxl no está instalado. Instala con: pip install openpyxl")

    import tempfile, shutil, urllib.parse

    def _file_uri_to_path(uri_or_path: str) -> str:
        s = (uri_or_path or "").strip()
        if s.lower().startswith("file:///"):
            return urllib.parse.unquote(s.replace("file:///", ""))
        return s

    def _persist_temp_copy(src_path: str, suffix: str = ".png") -> str:
        """Crea copia temporal PERSISTENTE y cierra el handle (Windows-safe)."""
        if not src_path or not os.path.exists(src_path):
            return ""
        fd, dst = tempfile.mkstemp(prefix="facot_", suffix=suffix, dir=tempfile.gettempdir())
        os.close(fd)
        shutil.copy2(src_path, dst)
        return dst

    template = _resolve_template(invoice_data or {}, template)

    wb = Workbook()
    sh = wb.active
    sh.title = "Factura"

    # Estilos
    bold = Font(bold=True)
    header_fill = PatternFill("solid", fgColor="E8EEF6")
    border = Border(left=Side(style="thin"), right=Side(style="thin"),
                    top=Side(style="thin"), bottom=Side(style="thin"))

    # Logo (robusto)
    tmp_files: List[str] = []
    try:
        if template.get("show_logo"):
            raw_logo = template.get("logo_path") or invoice_data.get("company_logo") or ""
            raw_logo = _file_uri_to_path(raw_logo)
            if raw_logo:
                logo_abs = _resolve_logo_path_rel_to_abs(raw_logo) or raw_logo
                if logo_abs and os.path.exists(logo_abs) and XLImage is not None:
                    logo_w = int(template.get("logo_width_px", 180) or 180)
                    try:
                        if PILImage is not None:
                            pil = PILImage.open(logo_abs)
                            ratio = (pil.height / pil.width) if pil.width else 1.0
                            new_h = max(1, int(logo_w * ratio))
                            # Copia temporal persistente
                            tmp_png = _persist_temp_copy(logo_abs, ".png")
                            if tmp_png:
                                try:
                                    pil_resized = pil.resize((logo_w, new_h))
                                    pil_resized.save(tmp_png, format="PNG")
                                except Exception:
                                    pass
                                tmp_files.append(tmp_png)
                                img = XLImage(tmp_png)
                                sh.add_image(img, "A1")
                        else:
                            tmp_cp = _persist_temp_copy(logo_abs, os.path.splitext(logo_abs)[1] or ".png")
                            if tmp_cp:
                                tmp_files.append(tmp_cp)
                                img = XLImage(tmp_cp)
                                sh.add_image(img, "A1")
                    except Exception:
                        # fallback: intentar con el original
                        try:
                            img = XLImage(logo_abs)
                            sh.add_image(img, "A1")
                        except Exception:
                            pass
    except Exception:
        pass

    # Encabezado
    sh.merge_cells("A1:H1")
    sh["A1"] = company_name or invoice_data.get("company_name", "")
    sh["A1"].font = Font(size=14, bold=True)

    sh["A2"] = "Tipo NCF:"
    sh["B2"] = invoice_data.get("ncf_type", "")
    sh["C2"] = "NCF #:"
    sh["D2"] = invoice_data.get("ncf_number", "")

    sh["A3"] = "Fecha:"
    sh["B3"] = invoice_data.get("invoice_date", datetime.date.today().isoformat())
    sh["C3"] = "Vencimiento:"
    sh["D3"] = invoice_data.get("due_date", "")

    sh["A4"] = "Cliente:"
    sh["B4"] = invoice_data.get("client_name", "")
    sh["C4"] = "RNC/Cédula:"
    sh["D4"] = invoice_data.get("client_rnc", "")

    sh["A5"] = "Moneda:"
    sh["B5"] = invoice_data.get("currency", "RD$")
    apply_itbis = bool(invoice_data.get("apply_itbis", False))
    itbis_rate = float(invoice_data.get("itbis_rate", 0.0) or 0.0)

    # Tabla de items
    start_row = 8
    headers = ["#", "Código", "Descripción", "Unidad", "Cantidad", "Precio Unit.", "Descuento %", "Subtotal"]
    for idx, h in enumerate(headers, start=1):
        c = sh.cell(row=start_row, column=idx, value=h)
        c.font = bold
        # Color primario si existe
        try:
            primary = template.get("primary_color")
            if primary:
                c.fill = PatternFill("solid", fgColor=primary.replace("#", ""))
            else:
                c.fill = header_fill
        except Exception:
            c.fill = header_fill
        c.alignment = Alignment(horizontal="center")
        c.border = border

    row = start_row + 1
    total = 0.0
    for i, it in enumerate(items, start=1):
        code = it.get("code", "") or it.get("item_code", "")
        desc = it.get("description", "") or it.get("nombre", "")
        unit = it.get("unit", "") or it.get("unidad", "")
        try:
            qty = float(it.get("quantity", it.get("cantidad", 0)) or 0)
        except Exception:
            qty = 0.0
        try:
            price = float(it.get("unit_price", it.get("precio", 0)) or 0)
        except Exception:
            price = 0.0
        disc = float(it.get("discount_pct", 0) or 0)  # %

        line_total = qty * price * (1 - disc / 100.0)
        total += line_total

        sh.cell(row=row, column=1, value=i).border = border
        sh.cell(row=row, column=2, value=code).border = border
        sh.cell(row=row, column=3, value=desc).border = border
        sh.cell(row=row, column=4, value=unit).border = border

        qcell = sh.cell(row=row, column=5, value=qty)
        qcell.number_format = '#,##0.00'  # SIEMPRE 2 decimales
        qcell.alignment = Alignment(horizontal="right")
        qcell.border = border

        pcell = sh.cell(row=row, column=6, value=price)
        pcell.number_format = '#,##0.00'
        pcell.alignment = Alignment(horizontal="right")
        pcell.border = border

        dcell = sh.cell(row=row, column=7, value=disc)
        dcell.number_format = '0.00'
        dcell.alignment = Alignment(horizontal="right")
        dcell.border = border

        scell = sh.cell(row=row, column=8, value=line_total)
        scell.number_format = '#,##0.00'
        scell.alignment = Alignment(horizontal="right")
        scell.border = border

        row += 1

    # Totales (alineados a la derecha)
    lbl_sub = sh.cell(row=row + 1, column=7, value="Subtotal:")
    lbl_sub.font = bold
    lbl_sub.alignment = Alignment(horizontal="right")
    val_sub = sh.cell(row=row + 1, column=8, value=total)
    val_sub.number_format = '#,##0.00'
    val_sub.alignment = Alignment(horizontal="right")

    itbis = total * itbis_rate if apply_itbis else 0.0
    if itbis > 0:
        lbl_itb = sh.cell(row=row + 2, column=7, value=f"ITBIS ({itbis_rate*100:.0f}%):")
        lbl_itb.font = bold
        lbl_itb.alignment = Alignment(horizontal="right")
        val_itb = sh.cell(row=row + 2, column=8, value=itbis)
        val_itb.number_format = '#,##0.00'
        val_itb.alignment = Alignment(horizontal="right")

        lbl_tot = sh.cell(row=row + 3, column=7, value="Total:")
        lbl_tot.font = bold
        lbl_tot.alignment = Alignment(horizontal="right")
        val_tot = sh.cell(row=row + 3, column=8, value=total + itbis)
        val_tot.number_format = '#,##0.00'
        val_tot.alignment = Alignment(horizontal="right")
    else:
        lbl_tot = sh.cell(row=row + 2, column=7, value="Total:")
        lbl_tot.font = bold
        lbl_tot.alignment = Alignment(horizontal="right")
        val_tot = sh.cell(row=row + 2, column=8, value=total)
        val_tot.number_format = '#,##0.00'
        val_tot.alignment = Alignment(horizontal="right")

    # Notas
    notes = invoice_data.get("notes", "")
    if notes:
        nr = row + 5
        sh.cell(row=nr, column=1, value="Notas:")
        sh.cell(row=nr + 1, column=1, value=notes)

    # Anchos de columna
    widths = [6, 14, 50, 10, 12, 18, 12, 18]
    for i, w in enumerate(widths, start=1):
        try:
            sh.column_dimensions[get_column_letter(i)].width = w
        except Exception:
            pass

    # Guardar y limpiar temporales
    try:
        wb.save(save_path)
    finally:
        for tf in tmp_files:
            try:
                if tf and os.path.exists(tf):
                    os.remove(tf)
            except Exception:
                pass
            
def generate_invoice_pdf(invoice_data: Dict, items: List[Dict], save_path: str, company_name: str = "", template: Dict = None):
    """
    Genera un PDF de la Factura. Requiere reportlab.
    Acepta template (dict) para logo/encabezado/colores.
    """
    if not REPORTLAB_AVAILABLE:
        raise RuntimeError("reportlab no está instalado. Instala con: pip install reportlab")

    template = _resolve_template(invoice_data or {}, template)
    doc = SimpleDocTemplate(save_path, pagesize=A4, rightMargin=24, leftMargin=24, topMargin=24, bottomMargin=24)
    styles = getSampleStyleSheet()
    elements = []

    # Logo
    logo_path_rel = template.get("logo_path") or invoice_data.get("company_logo", "")
    if logo_path_rel:
        logo_abs = _resolve_logo_path_rel_to_abs(logo_path_rel)
        if logo_abs:
            try:
                logo_w = float(template.get("logo_width_px", 120) or 120)
                # RLImage auto-scales height if you pass width and None for height is not supported; pass height based on image ratio
                try:
                    if PILImage is not None:
                        pil = PILImage.open(logo_abs)
                        ratio = pil.height / pil.width if pil.width else 1.0
                        logo_h = int(logo_w * ratio)
                        rlimg = RLImage(logo_abs, width=logo_w, height=logo_h)
                    else:
                        rlimg = RLImage(logo_abs, width=logo_w, height=None)
                    elements.append(rlimg)
                    elements.append(Spacer(1, 6))
                except Exception:
                    # fallback: try ImageReader usage or skip
                    try:
                        img_reader = ImageReader(logo_abs)
                        # skipping detailed placement, append as paragraph with image may be more complex
                    except Exception:
                        pass
            except Exception:
                pass

    # Header lines from template
    for line in template.get("header_lines", []) or []:
        if line and line.strip():
            elements.append(Paragraph(line, styles["Normal"]))
    elements.append(Spacer(1, 8))

    # Client/meta info
    meta = f"Fecha: {invoice_data.get('invoice_date', datetime.date.today().isoformat())}    Moneda: {invoice_data.get('currency','RD$')}"
    ncf_line = f"Tipo NCF: {invoice_data.get('ncf_type','')}   NCF #: {invoice_data.get('ncf_number','')}"
    elements.append(Paragraph(ncf_line, styles["Normal"]))
    elements.append(Paragraph(meta, styles["Normal"]))
    elements.append(Spacer(1, 12))
    client = f"Cliente: {invoice_data.get('client_name','')}  -  RNC: {invoice_data.get('client_rnc','')}"
    elements.append(Paragraph(client, styles["Normal"]))
    elements.append(Spacer(1, 12))

    # Table of items
    table_data = [["#", "Código", "Descripción", "Unidad", "Cantidad", "Precio Unit.", "Desc %", "Subtotal"]]
    total = 0.0
    for i, it in enumerate(items, start=1):
        code = it.get("code", "")
        desc = it.get("description", "") or it.get("nombre", "")
        unit = it.get("unit", "") or it.get("unidad", "")
        qty = float(it.get("quantity", it.get("cantidad", 0)) or 0)
        price = float(it.get("unit_price", it.get("precio", 0)) or 0)
        disc = float(it.get("discount_pct", 0) or 0)
        subtotal = qty * price * (1 - disc / 100.0)
        total += subtotal
        table_data.append([str(i), code, desc, unit, f"{qty:.2f}", f"{price:,.2f}", f"{disc:.2f}", f"{subtotal:,.2f}"])

    t = Table(table_data, colWidths=[30, 60, 200, 40, 50, 70, 50, 80])
    t.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#E8EEF6")),
        ('GRID', (0, 0), (-1, -1), 0.25, colors.gray),
        ('ALIGN', (4, 1), (7, -1), 'RIGHT'),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
    ]))
    elements.append(t)
    elements.append(Spacer(1, 12))

    apply_itbis = bool(invoice_data.get("apply_itbis", False))
    itbis_rate = float(invoice_data.get("itbis_rate", 0.0) or 0.0)
    itbis = total * itbis_rate if apply_itbis else 0.0

    elements.append(Paragraph(f"Subtotal: {total:,.2f}", styles["Normal"]))
    if itbis:
        elements.append(Paragraph(f"ITBIS ({itbis_rate*100:.0f}%): {itbis:,.2f}", styles["Normal"]))
    elements.append(Paragraph(f"Total: {total + itbis:,.2f}", styles["Normal"]))
    elements.append(Spacer(1, 12))

    notes = invoice_data.get("notes", "")
    if notes:
        elements.append(Paragraph("Notas:", styles["Heading4"]))
        elements.append(Paragraph(notes, styles["Normal"]))

    doc.build(elements)